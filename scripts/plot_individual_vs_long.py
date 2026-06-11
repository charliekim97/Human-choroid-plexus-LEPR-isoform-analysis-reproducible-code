#!/usr/bin/env python3
"""Per-transcript (individual, NOT summed) LepRa vs LepRb comparison (PI request):
   (1) ENST00000371060 (LEPR-205) vs ENST00000349533 (LEPR-202, long)
   (2) ENST00000616738 (LEPR-208) vs ENST00000349533
   control donors, per cohort, paired Wilcoxon signed-rank; + combined controls.
   No re-run: recomputed from existing per-transcript Salmon TPM."""
import csv, os, numpy as np
from scipy.stats import wilcoxon
import matplotlib as mpl; mpl.use("Agg"); import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT="/sessions/ecstatic-funny-pasteur/mnt/LEPR_human_ChP_RNAseq"
RES, FIG = ROOT+"/results", ROOT+"/figures"
mpl.rcParams.update({"font.family":"sans-serif","font.sans-serif":["Arial","DejaVu Sans"],
 "font.size":8,"axes.linewidth":0.8,"pdf.fonttype":42,"ps.fonttype":42,"svg.fonttype":"none"})

C={"GSE228458":dict(f="salmon_lepr_tpm.csv",color="#4C72B0"),
   "GSE137619":dict(f="LEPR_expression_GSE137619.csv",color="#DD8452")}
LONG="LEPR-202_ENST00000349533_TPM"
SHORTS=[("ENST00000371060","LEPR-205_ENST00000371060_TPM","LepRa  ENST00000371060 (LEPR-205)"),
        ("ENST00000616738","LEPR-208_ENST00000616738_TPM","LepRa  ENST00000616738 (LEPR-208)")]

def load_ctrl(fn):
    rows=[r for r in csv.DictReader(open(os.path.join(RES,fn))) if r["group"]=="control"]
    return rows

DATA={coh:load_ctrl(c["f"]) for coh,c in C.items()}

def stat(short,long):
    short=np.array(short); long=np.array(long)
    n=len(short); ngt=int(np.sum(short>long))
    # paired Wilcoxon (two-sided); guard zero-difference
    try: _,p=wilcoxon(short,long,alternative="two-sided")
    except Exception: p=float("nan")
    with np.errstate(divide="ignore",invalid="ignore"):
        fold=np.median(short/long)
    return n,ngt,p,fold

# ---- console + collect ----
print("="*72)
results={}
for enst,col,label in SHORTS:
    print(f"\n### {label}  vs  LEPR-202 ENST00000349533 (long)")
    alls=[]; alll=[]
    for coh in C:
        s=[float(r[col]) for r in DATA[coh]]; l=[float(r[LONG]) for r in DATA[coh]]
        n,ngt,p,fold=stat(s,l); alls+=s; alll+=l
        results[(enst,coh)]=dict(s=s,l=l,n=n,ngt=ngt,p=p,fold=fold)
        print(f"  {coh:10s} n={n}  short>long {ngt}/{n}  median fold={fold:.1f}x  paired Wilcoxon P={p:.4f}")
    nC,ngtC,pC,foldC=stat(alls,alll)
    results[(enst,"combined")]=dict(n=nC,ngt=ngtC,p=pC,fold=foldC)
    print(f"  COMBINED   n={nC}  short>long {ngtC}/{nC}  median fold={foldC:.1f}x  paired Wilcoxon P={pC:.4f}")

# ---- FIGURE: 2 transcripts (rows) x 2 cohorts (cols), paired before-after, log y ----
fig,axes=plt.subplots(2,2,figsize=(8.4,8.0),gridspec_kw=dict(hspace=0.42,wspace=0.30))
for i,(enst,col,label) in enumerate(SHORTS):
    for j,coh in enumerate(C):
        ax=axes[i,j]; c=C[coh]["color"]; R=results[(enst,coh)]
        s=np.array(R["s"]); l=np.array(R["l"]); floor=1e-3
        sp=np.clip(s,floor,None); lp=np.clip(l,floor,None)
        for k in range(len(s)):
            ax.plot([0,1],[sp[k],lp[k]],color="#999",lw=0.8,zorder=1)
        ax.scatter(np.zeros(len(s)),sp,s=34,c=c,edgecolor="white",lw=0.5,zorder=3)
        ax.scatter(np.ones(len(l)),lp,s=34,c="#9aa0a6",edgecolor="white",lw=0.5,zorder=3)
        ax.set_yscale("log"); ax.set_xlim(-0.4,1.4); ax.set_xticks([0,1])
        ax.set_xticklabels([enst.replace("ENST00000","…"),"ENST…349533\n(long)"],fontsize=7.5)
        ax.set_ylabel("TPM (log)")
        ax.set_title(f"{coh}  (control n={R['n']})",fontsize=9)
        ax.text(0.5,1.18,f"short>long {R['ngt']}/{R['n']} · {R['fold']:.0f}× · P={R['p']:.3f}",
                transform=ax.transAxes,ha="center",va="top",fontsize=7.6,color="#333")
        ax.spines[["top","right"]].set_visible(False)
    axes[i,0].annotate(label,xy=(-0.42,0.5),xycoords="axes fraction",rotation=90,
                       ha="center",va="center",fontsize=9.5,fontweight="bold",color="#22304A")
cmb1=results[("ENST00000371060","combined")]; cmb2=results[("ENST00000616738","combined")]
fig.suptitle("Each LepRa transcript individually exceeds LepRb (long) in control human ChP\n"
   f"combined controls (n=11):  371060 vs long P={cmb1['p']:.3f} ({cmb1['ngt']}/11) · "
   f"616738 vs long P={cmb2['p']:.3f} ({cmb2['ngt']}/11)",
   fontsize=10.5,fontweight="bold",y=1.0)
for ext in ("png","pdf","svg"): fig.savefig(f"{FIG}/Fig_LEPR_individual_vs_long.{ext}",dpi=300,bbox_inches="tight")
plt.close(fig)

# ---- per-donor CSV ----
with open(RES+"/LEPR_individual_vs_long_perdonor.csv","w",newline="") as f:
    w=csv.writer(f); w.writerow(["cohort","donor","ENST00000371060_TPM","ENST00000616738_TPM","ENST00000349533_long_TPM"])
    for coh in C:
        for r in DATA[coh]:
            w.writerow([coh,r["donor"],r["LEPR-205_ENST00000371060_TPM"],
                        r["LEPR-208_ENST00000616738_TPM"],r[LONG]])

# ---- Prism xlsx ----
HEAD=Font(bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="305496")
SUB=Font(bold=True,color="1F3864"); CEN=Alignment(horizontal="center")
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=Workbook(); ws=wb.active; ws.title="HOWTO"; ws.column_dimensions["A"].width=104
howto=[("개별 transcript vs long — Prism용 (합산 안 함)",True),
 ("",False),
 ("비교 1: ENST00000371060 (LEPR-205)  vs  ENST00000349533 (LEPR-202, long)",True),
 ("비교 2: ENST00000616738 (LEPR-208)  vs  ENST00000349533 (LEPR-202, long)",True),
 ("",False),
 ("각 시트에서 cohort별로 [short transcript TPM] 와 [long TPM] 두 열을 Prism에 붙여넣기 (행=donor).",False),
 ("Before-after(paired) 그래프 · Y축 Log10 · Wilcoxon matched-pairs signed rank test. cohort는 합치지 않고 각각 1장.",False),
 ("",False),
 (f"통계 요약 (control): 371060 vs long → GSE228458 {results[('ENST00000371060','GSE228458')]['ngt']}/5 P={results[('ENST00000371060','GSE228458')]['p']:.4f}, "
  f"GSE137619 {results[('ENST00000371060','GSE137619')]['ngt']}/6 P={results[('ENST00000371060','GSE137619')]['p']:.4f}, "
  f"combined 11 P={cmb1['p']:.3f}",False),
 (f"            616738 vs long → GSE228458 {results[('ENST00000616738','GSE228458')]['ngt']}/5 P={results[('ENST00000616738','GSE228458')]['p']:.4f}, "
  f"GSE137619 {results[('ENST00000616738','GSE137619')]['ngt']}/6 P={results[('ENST00000616738','GSE137619')]['p']:.4f}, "
  f"combined 11 P={cmb2['p']:.3f}",False),
]
for i,(t,b) in enumerate(howto,1):
    ws.cell(i,1,t).font=Font(bold=b)

def sheet(name,col,title):
    ws=wb.create_sheet(name)
    for cc,wd in zip("ABC",[10,22,22]): ws.column_dimensions[cc].width=wd
    r=1
    for coh in C:
        ws.cell(r,1,f"{coh} control  ({title} vs long, log Y, Wilcoxon paired)").font=SUB; r+=1
        for j,h in enumerate(["Donor",title+"_TPM","ENST349533_long_TPM"],1):
            c=ws.cell(r,j,h); c.font=HEAD; c.fill=HF; c.alignment=CEN; c.border=BORD
        r+=1
        for x in DATA[coh]:
            ws.cell(r,1,x["donor"]).border=BORD
            ws.cell(r,2,round(float(x[col]),4)).border=BORD
            ws.cell(r,3,round(float(x[LONG]),4)).border=BORD
            r+=1
        r+=1
sheet("t371060_vs_long","LEPR-205_ENST00000371060_TPM","ENST371060")
sheet("t616738_vs_long","LEPR-208_ENST00000616738_TPM","ENST616738")
wb.save(ROOT+"/For_PI/LEPR_Prism_individual_vs_long.xlsx")
print("\nwrote Fig_LEPR_individual_vs_long.{png,pdf,svg}, LEPR_individual_vs_long_perdonor.csv, For_PI/LEPR_Prism_individual_vs_long.xlsx")
